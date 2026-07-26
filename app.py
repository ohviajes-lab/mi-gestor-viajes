import streamlit as st
import pandas as pd
import json
import io
import sqlite3
from PIL import Image
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

# --- BASE DE DATOS (SQLITE) ---
DB_NAME = "presupuestos.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS presupuestos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            destino TEXT,
            fecha_salida TEXT,
            duracion TEXT,
            operador TEXT,
            total_usd REAL,
            datos_json TEXT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def guardar_en_db(datos, total_calc):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
        INSERT INTO presupuestos (destino, fecha_salida, duracion, operador, total_usd, datos_json)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (datos['destino'], datos['fecha_salida'], datos['duracion'], datos['operador'], total_calc, json.dumps(datos)))
    conn.commit()
    conn.close()

def cargar_historial():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query("SELECT id, destino, fecha_salida, operador, total_usd, fecha_registro FROM presupuestos ORDER BY id DESC", conn)
    conn.close()
    return df

# Inicializar Base de Datos
init_db()

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Gestor de Presupuestos de Viaje", page_icon="✈️", layout="wide")

st.title("✈️ Consolidador de Presupuestos de Viaje")

# Opciones de Navegación
menu = st.sidebar.radio("Navegación", ["🧮 Crear Presupuesto", "📜 Historial Guardado"])

st.sidebar.markdown("---")
st.sidebar.header("⚙️ Configuración")
api_key = st.sidebar.text_input("Ingresa tu Gemini API Key", type="password", help="Obtenla gratis en Google AI Studio")

# Session State inicial
if "form_version" not in st.session_state:
# --- DICCIONARIO INICIAL Y CONTROL DE ESTADO ---
if "form_version" not in st.session_state:
    st.session_state["form_version"] = 0

if "datos_viaje" not in st.session_state:
    st.session_state["datos_viaje"] = {
        # ... (tus datos por defecto) ...
    }
if "datos_viaje" not in st.session_state:
    st.session_state["datos_viaje"] = {
        "destino": "Costa Rica",
        "fecha_salida": "04/12/2026",
        "duracion": "9 Días / 8 Noches",
        "operador": "Mediterránea Turismo",
        "precio_base": 2649.00,
        "impuestos_aereo": 444.00,
        "cuotas_cant": 5,
        "cuota_monto": 618.60,
        "equipaje_extra": 100.00,
        "comidas_extra": 280.00,
        "excursiones_extra": 120.00,
        "seguro_medico": 75.00,
        "gastos_personales": 220.00,
        "notas": "Incluye 6N c/desayuno + 2N pensión completa + traslados y tours."
    }

# --- GENERADORES DE ARCHIVOS ---
def generar_excel(datos, total_calc):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto Viaje"
    
    font_title = Font(name="Calibri", size=14, bold=True, color="1E4D2B")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    fill_header = PatternFill(start_color="1E4D2B", end_color="1E4D2B", fill_type="solid")
    fill_accent = PatternFill(start_color="E8F0E6", end_color="E8F0E6", fill_type="solid")
    
    ws["A1"] = f"PRESUPUESTO INTEGRADO - {datos['destino'].upper()}"
    ws["A1"].font = font_title
    ws["A2"] = f"Fecha: {datos['fecha_salida']} | Duración: {datos['duracion']} | Operador: {datos['operador']}"
    
    headers = ["Categoría", "Concepto / Ítem", "Fuente", "Costo (USD)"]
    ws.append([])
    ws.append(headers)
    
    for col_num in range(1, 5):
        cell = ws.cell(row=4, column=col_num)
        cell.font = font_header
        cell.fill = fill_header

    filas = [
        ("Transporte", "Paquete Base Terrestre", "Imagen / Promo", datos['precio_base']),
        ("Transporte", "Impuestos Aéreos", "Imagen / Promo", datos['impuestos_aereo']),
        ("Transporte", "Equipaje Facturado Extra", "Manual", datos['equipaje_extra']),
        ("Alimentación", "Comidas No Incluidas", "Manual", datos['comidas_extra']),
        ("Excursiones", "Actividades Extras / Parques", "Manual", datos['excursiones_extra']),
        ("Seguro", "Asistencia Médica Internacional", "Manual", datos['seguro_medico']),
        ("Varios", "Propinas e Imprevistos", "Manual", datos['gastos_personales']),
    ]
    
    for f in filas:
        ws.append(list(f))
        
    tot_row = len(filas) + 5
    ws.cell(row=tot_row, column=1, value="TOTAL ESTIMADO (USD)").font = font_bold
    ws.cell(row=tot_row, column=4, value=f"=SUM(D5:D{tot_row-1})").font = font_bold
    
    for c in range(1, 5):
        ws.cell(row=tot_row, column=c).fill = fill_accent

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generar_pdf(datos, total_calc):
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Helvetica, Arial, sans-serif; color: #2C3E50; padding: 20px; }}
            .header {{ background-color: #1E4D2B; color: white; padding: 15px; border-radius: 6px; }}
            h1 {{ margin: 0; font-size: 20px; }}
            .card {{ background-color: #F8FAF7; border: 1px solid #D3D3D3; padding: 10px; margin-top: 15px; border-radius: 6px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th {{ background-color: #1E4D2B; color: white; padding: 8px; font-size: 12px; }}
            td {{ padding: 8px; border-bottom: 1px solid #E0E0E0; font-size: 12px; }}
            .total {{ font-weight: bold; background-color: #E8F0E6; }}
            .text-right {{ text-align: right; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Presupuesto Integrado de Viaje: {datos['destino']}</h1>
            <p>Salida: {datos['fecha_salida']} | Duración: {datos['duracion']} | Operador: {datos['operador']}</p>
        </div>
        
        <div class="card">
            <h3>Resumen Económico Total: USD ${total_calc:,.2f}</h3>
            <p><strong>Financiación Base:</strong> {datos['cuotas_cant']} cuotas de USD ${datos['cuota_monto']:,.2f}</p>
        </div>

        <table>
            <thead>
                <tr>
                    <th>Categoría</th>
                    <th>Concepto</th>
                    <th>Fuente</th>
                    <th class="text-right">Monto (USD)</th>
                </tr>
            </thead>
            <tbody>
                <tr><td>Transporte</td><td>Paquete Base Terrestre</td><td>Imagen / Promo</td><td class="text-right">${datos['precio_base']:,.2f}</td></tr>
                <tr><td>Transporte</td><td>Impuestos Aéreos</td><td>Imagen / Promo</td><td class="text-right">${datos['impuestos_aereo']:,.2f}</td></tr>
                <tr><td>Transporte</td><td>Equipaje Facturado Extra</td><td>Estimación Manual</td><td class="text-right">${datos['equipaje_extra']:,.2f}</td></tr>
                <tr><td>Alimentación</td><td>Comidas No Incluidas</td><td>Estimación Manual</td><td class="text-right">${datos['comidas_extra']:,.2f}</td></tr>
                <tr><td>Excursiones</td><td>Tours Extra / Entradas</td><td>Estimación Manual</td><td class="text-right">${datos['excursiones_extra']:,.2f}</td></tr>
                <tr><td>Seguro</td><td>Asistencia Médica Internacional</td><td>Estimación Manual</td><td class="text-right">${datos['seguro_medico']:,.2f}</td></tr>
                <tr><td>Varios</td><td>Propinas e Imprevistos</td><td>Estimación Manual</td><td class="text-right">${datos['gastos_personales']:,.2f}</td></tr>
                <tr class="total">
                    <td colspan="3"><strong>TOTAL ESTIMADO POR PERSONA</strong></td>
                    <td class="text-right"><strong>${total_calc:,.2f}</strong></td>
                </tr>
            </tbody>
        </table>
        
        <div class="card" style="margin-top:20px;">
            <p><strong>Notas adicionales:</strong> {datos['notas']}</p>
        </div>
    </body>
    </html>
    """
    output = io.BytesIO()
    pisa.CreatePDF(html_content, dest=output)
    return output.getvalue()


# --- VISTA 1: CREAR PRESUPUESTO ---
if menu == "🧮 Crear Presupuesto":
    st.subheader("1. Carga de Documentos y Fuentes de Información")
    col_left, col_right = st.columns(2)

    with col_left:
        uploaded_image = st.file_uploader("🖼️ Imagen / Flyer / Comprobante (JPG, PNG)", type=["jpg", "png", "jpeg"])
        web_url = st.text_input("🌐 Enlace Web (opcional)")

    with col_right:
        email_text = st.text_area("📧 Texto de Correo Electrónico / Notas", height=140, placeholder="Pega aquí el contenido del mail de la agencia...")

    if st.button("🔍 Extraer Datos con IA (Gemini)"):
        if not api_key:
            st.error("⚠️ Por favor ingresa tu API Key de Gemini en el panel lateral.")
        elif not (uploaded_image or email_text or web_url):
            st.warning("⚠️ Debes proporcionar al menos una fuente de datos.")
        else:
            with st.spinner("Analizando información con IA..."):
                try:
                    client = genai.Client(api_key=api_key)
                    contents = []
                    if uploaded_image:
                        contents.append(Image.open(uploaded_image))
                    
                    prompt = f"""
                    Analiza la información de viaje y responde ÚNICAMENTE con un JSON válido con la estructura:
                    {{
                        "destino": "Nombre del destino", "fecha_salida": "Fecha o mes", "duracion": "Días/noches",
                        "operador": "Agencia/Operador", "precio_base": float, "impuestos_aereo": float,
                        "cuotas_cant": int, "cuota_monto": float, "notas": "Resumen"
                    }}
                    Si no hay dato, usa 0.0. Web: {web_url} | Mail: {email_text}
                    """
                    contents.append(prompt)

                    response = client.models.generate_content(
                        model="gemini-2.5-flash",
                        contents=contents,
                        config=types.GenerateContentConfig(response_mime_type="application/json")
                    )

                    extracted_data = json.loads(response.text)
                    for key, val in extracted_data.items():
                        if val is not None and val != 0:
                            st.session_state["datos_viaje"][key] = val

                    st.success("✨ ¡Datos extraídos con éxito!")
# Actualizar los datos extraídos
                for key, val in extracted_data.items():
                    if val is not None and val != 0:
                        st.session_state["datos_viaje"][key] = val

                # 💡 AGREGAR ESTAS 2 LÍNEAS AQUÍ:
                st.session_state["form_version"] += 1
                st.rerun()

            except Exception as e:
                st.error(f"Error: {str(e)}")

                except Exception as e:
                    st.error(f"Error: {str(e)}")

    st.markdown("---")
    st.subheader("2. Edición y Completado Manual de Presupuesto")
v = st.session_state["form_version"]

    with st.form("form_presupuesto"):
        st.markdown("#### 📌 Datos Generales del Viaje")
        c1, c2, c3, c4 = st.columns(4)
        with c1: destino = st.text_input("Destino", value=st.session_state["datos_viaje"]["destino"], key=f"dest_{v}")
        with c2: fecha_salida = st.text_input("Fecha de Salida", value=st.session_state["datos_viaje"]["fecha_salida"], key=f"fec_{v}")
        with c3: duracion = st.text_input("Duración", value=st.session_state["datos_viaje"]["dur_{v}"] if "dur_{v}" in st.session_state else st.session_state["datos_viaje"]["duracion"], key=f"dur_{v}")
        with c4: operador = st.text_input("Operador / Agencia", value=st.session_state["datos_viaje"]["operador"], key=f"ope_{v}")

        st.markdown("#### 💳 Costos Confirmados")
        col_a, col_b, col_c, col_d = st.columns(4)
        with col_a: precio_base = st.number_input("Paquete Base (USD)", value=float(st.session_state["datos_viaje"]["precio_base"]), key=f"pbase_{v}")
        with col_b: impuestos_aereo = st.number_input("Impuestos Aéreos (USD)", value=float(st.session_state["datos_viaje"]["impuestos_aereo"]), key=f"imp_{v}")
        with col_c: cuotas_cant = st.number_input("Cantidad de Cuotas", value=int(st.session_state["datos_viaje"]["cuotas_cant"]), key=f"ccant_{v}")
        with col_d: cuota_monto = st.number_input("Monto Cuota (USD)", value=float(st.session_state["datos_viaje"]["cuota_monto"]), key=f"cmont_{v}")

        st.markdown("#### ✍️ Estimaciones Manuales")
        col_e, col_f, col_g, col_h = st.columns(4)
        with col_e: equipaje_extra = st.number_input("Equipaje Facturado (USD)", value=float(st.session_state["datos_viaje"]["equipaje_extra"]), key=f"eq_{v}")
        with col_f: comidas_extra = st.number_input("Comidas no incluidas (USD)", value=float(st.session_state["datos_viaje"]["comidas_extra"]), key=f"com_{v}")
        with col_g: excursiones_extra = st.number_input("Tours Extra (USD)", value=float(st.session_state["datos_viaje"]["excursiones_extra"]), key=f"exc_{v}")
        with col_h: seguro_medico = st.number_input("Seguro Médico (USD)", value=float(st.session_state["datos_viaje"]["seguro_medico"]), key=f"seg_{v}")

        gastos_personales = st.number_input("Propinas e Imprevistos (USD)", value=float(st.session_state["datos_viaje"]["gastos_personales"]), key=f"gpers_{v}")
        notas = st.text_area("Observaciones / Servicios Incluidos", value=st.session_state["datos_viaje"]["notas"], key=f"not_{v}")

        btn_guardar = st.form_submit_button("🧮 Recalcular Presupuesto y Preparar Descargas")

    with st.form("form_presupuesto"):
        st.markdown("#### 📌 Datos Generales del Viaje")
        c1, c2, c3, c4 = st.columns(4)
        with c1: destino = st.text_input("Destino", value=st.session_state["datos_viaje"]["destino"])
        with c2: fecha_salida = st.text_input("Fecha de Salida", value=st.session_state["datos_viaje"]["fecha_salida"])
        with c3: duracion = st.text_input("Duración", value=st.session_state["datos_viaje"]["duracion"])
        with c4: operador = st.text_input("Operador / Agencia", value=st.session_state["datos_viaje"]["operador"])

        st.markdown("#### 💳 Costos Confirmados")
        col_a, col_b, col_c, col_d = st.columns(4)
        with col_a: precio_base = st.number_input("Paquete Base (USD)", value=float(st.session_state["datos_viaje"]["precio_base"]))
        with col_b: impuestos_aereo = st.number_input("Impuestos Aéreos (USD)", value=float(st.session_state["datos_viaje"]["impuestos_aereo"]))
        with col_c: cuotas_cant = st.number_input("Cantidad de Cuotas", value=int(st.session_state["datos_viaje"]["cuotas_cant"]))
        with col_d: cuota_monto = st.number_input("Monto Cuota (USD)", value=float(st.session_state["datos_viaje"]["cuota_monto"]))

        st.markdown("#### ✍️ Estimaciones Manuales")
        col_e, col_f, col_g, col_h = st.columns(4)
        with col_e: equipaje_extra = st.number_input("Equipaje Facturado (USD)", value=float(st.session_state["datos_viaje"]["equipaje_extra"]))
        with col_f: comidas_extra = st.number_input("Comidas no incluidas (USD)", value=float(st.session_state["datos_viaje"]["comidas_extra"]))
        with col_g: excursiones_extra = st.number_input("Tours Extra (USD)", value=float(st.session_state["datos_viaje"]["excursiones_extra"]))
        with col_h: seguro_medico = st.number_input("Seguro Médico (USD)", value=float(st.session_state["datos_viaje"]["seguro_medico"]))

        gastos_personales = st.number_input("Propinas e Imprevistos (USD)", value=float(st.session_state["datos_viaje"]["gastos_personales"]))
        notas = st.text_area("Observaciones / Servicios Incluidos", value=st.session_state["datos_viaje"]["notas"])

        btn_guardar = st.form_submit_button("🧮 Recalcular Presupuesto y Preparar Descargas")

    st.markdown("---")
    st.subheader("3. Resumen Consolidado, Guardado y Descargas")

    datos_actuales = {
        "destino": destino, "fecha_salida": fecha_salida, "duracion": duracion, "operador": operador,
        "precio_base": precio_base, "impuestos_aereo": impuestos_aereo, "cuotas_cant": cuotas_cant,
        "cuota_monto": cuota_monto, "equipaje_extra": equipaje_extra, "comidas_extra": comidas_extra,
        "excursiones_extra": excursiones_extra, "seguro_medico": seguro_medico,
        "gastos_personales": gastos_personales, "notas": notas
    }

    total_calculado = (precio_base + impuestos_aereo + equipaje_extra + 
                       comidas_extra + excursiones_extra + seguro_medico + gastos_personales)

    m1, m2, m3 = st.columns(3)
    with m1: st.metric("Total Estimado del Viaje", f"USD ${total_calculado:,.2f}")
    with m2: st.metric("Paquete Base (Promo)", f"USD ${(precio_base + impuestos_aereo):,.2f}")
    with m3: st.metric("Gastos Extras (Manuales)", f"USD ${(total_calculado - (precio_base + impuestos_aereo)):,.2f}")

    # Botón de guardado en Base de Datos
    if st.button("💾 Guardar este Presupuesto en la Base de Datos"):
        guardar_en_db(datos_actuales, total_calculado)
        st.success("✅ ¡Presupuesto guardado correctamente en la Base de Datos!")

    st.markdown("---")
    col_dl1, col_dl2 = st.columns(2)
    excel_bytes = generar_excel(datos_actuales, total_calculado)
    pdf_bytes = generar_pdf(datos_actuales, total_calculado)

    with col_dl1:
        st.download_button(
            label="📊 Descargar Presupuesto en Excel (.xlsx)",
            data=excel_bytes,
            file_name=f"Presupuesto_{destino.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col_dl2:
        st.download_button(
            label="📄 Descargar Informe en PDF (.pdf)",
            data=pdf_bytes,
            file_name=f"Presupuesto_{destino.replace(' ', '_')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

# --- VISTA 2: HISTORIAL DE LA BASE DE DATOS ---
elif menu == "📜 Historial Guardado":
    st.subheader("📜 Presupuestos Guardados en la Base de Datos")
    df_historial = cargar_historial()
    
    if df_historial.empty:
        st.info("Aún no has guardado presupuestos en la base de datos.")
    else:
        st.dataframe(df_historial, use_container_width=True)