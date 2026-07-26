import json
import io
import sqlite3
import base64
import streamlit as st
import pandas as pd
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

# --- BASE DE DATOS (SQLITE) ---
DB_NAME = "presupuestos.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS presupuestos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            destino TEXT,
            fecha_salida TEXT,
            duracion TEXT,
            operador TEXT,
            total_usd REAL,
            datos_json TEXT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def guardar_en_db(datos, total_calc):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
        INSERT INTO presupuestos (destino, fecha_salida, duracion, operador, total_usd, datos_json)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (datos['destino'], datos['fecha_salida'], datos['duracion'], datos['operador'], total_calc, json.dumps(datos)))
    conn.commit()
    conn.close()

def cargar_historial():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query("SELECT id, destino, fecha_salida, operador, total_usd, fecha_registro FROM presupuestos ORDER BY id DESC", conn)
    conn.close()
    return df

init_db()

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Gestor de Presupuestos de Viaje", page_icon="✈️", layout="wide")
st.title("✈️ Consolidador de Presupuestos de Viaje")

menu = st.sidebar.radio("Navegación", ["🧮 Crear Presupuesto", "📜 Historial Guardado"])

st.sidebar.markdown("---")
st.sidebar.header("⚙️ Configuración")

# API KEY DE GROQ
api_key = st.secrets.get("GROQ_API_KEY", "")
if api_key:
    st.sidebar.success("🔑 API Key cargada desde Secrets")
    api_key = st.sidebar.text_input("Groq API Key", value=api_key, type="password")
else:
    api_key = st.sidebar.text_input("Ingresa tu Groq API Key (gsk_...)", type="password")

# SELECCIÓN DE MODELOS
st.sidebar.markdown("---")
st.sidebar.subheader("🤖 Selección de Modelos")

modelo_texto = st.sidebar.selectbox(
    "Modelo para Texto / Links:",
    ["llama-3.3-70b-versatile", "llama3-8b-8192", "mixtral-8x7b-32768"],
    index=0
)

modelo_vision = st.sidebar.selectbox(
    "Modelo para Imágenes:",
    ["llama-3.2-11b-vision-instruct", "llama-3.2-90b-vision-instruct"],
    index=0
)

# Estado de la sesión
if "form_version" not in st.session_state:
    st.session_state["form_version"] = 0

if "datos_viaje" not in st.session_state:
    st.session_state["datos_viaje"] = {
        "destino": "Costa Rica", "fecha_salida": "04/12/2026", "duracion": "9 Días / 8 Noches",
        "operador": "Mediterránea Turismo", "precio_base": 2649.00, "impuestos_aereo": 444.00,
        "cuotas_cant": 5, "cuota_monto": 618.60, "equipaje_extra": 100.00, "comidas_extra": 280.00,
        "excursiones_extra": 120.00, "seguro_medico": 75.00, "gastos_personales": 220.00,
        "notas": "Incluye 6N c/desayuno + 2N pensión completa + traslados y tours."
    }

# EXCEL Y PDF GENERADORES
def generar_excel(datos, total_calc):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto Viaje"
    ws["A1"] = f"PRESUPUESTO INTEGRADO - {datos['destino'].upper()}"
    ws["A2"] = f"Fecha: {datos['fecha_salida']} | Duración: {datos['duracion']} | Operador: {datos['operador']}"
    
    headers = ["Categoría", "Concepto / Ítem", "Fuente", "Costo (USD)"]
    ws.append([])
    ws.append(headers)

    filas = [
        ("Transporte", "Paquete Base Terrestre", "Imagen / Promo", datos['precio_base']),
        ("Transporte", "Impuestos Aéreos", "Imagen / Promo", datos['impuestos_aereo']),
        ("Transporte", "Equipaje Facturado Extra", "Manual", datos['equipaje_extra']),
        ("Alimentación", "Comidas No Incluidas", "Manual", datos['comidas_extra']),
        ("Excursiones", "Actividades Extras / Parques", "Manual", datos['excursiones_extra']),
        ("Seguro", "Asistencia Médica Internacional", "Manual", datos['seguro_medico']),
        ("Varios", "Propinas e Imprevistos", "Manual", datos['gastos_personales']),
    ]
    for f in filas:
        ws.append(list(f))
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generar_pdf(datos, total_calc):
    html_content = f"""
    <html>
    <body>
        <h2>Presupuesto Integrado de Viaje: {datos['destino']}</h2>
        <p><strong>Salida:</strong> {datos['fecha_salida']} | <strong>Duración:</strong> {datos['duracion']} | <strong>Operador:</strong> {datos['operador']}</p>
        <hr/>
        <h3>TOTAL ESTIMADO: USD ${total_calc:,.2f}</h3>
        <p><strong>Notas:</strong> {datos['notas']}</p>
    </body>
    </html>
    """
    output = io.BytesIO()
    pisa.CreatePDF(html_content, dest=output)
    return output.getvalue()

# VISTA 1: CREAR PRESUPUESTO
if menu == "🧮 Crear Presupuesto":
    st.subheader("1. Carga de Documentos y Fuentes de Información")
    col_left, col_right = st.columns(2)

    with col_left:
        uploaded_image = st.file_uploader("🖼️ Imagen / Flyer (JPG, PNG)", type=["jpg", "png", "jpeg"])
        web_url = st.text_input("🌐 Enlace Web (opcional)")

    with col_right:
        email_text = st.text_area("📧 Texto / Email de la Agencia", height=140)

    if st.button("🔍 Extraer Datos con IA"):
        if not api_key:
            st.error("⚠️ Falta la API Key de Groq.")
        elif not (uploaded_image or email_text or web_url):
            st.warning("⚠️ Sube una imagen o pega un texto/enlace.")
        else:
            # Enrutamiento automático
            modelo_a_usar = modelo_vision if uploaded_image else modelo_texto

            with st.spinner(f"Procesando con {modelo_a_usar}..."):
                try:
                    client = Groq(api_key=api_key)
                    prompt_text = (
                        "Analiza la información de viaje y responde ÚNICAMENTE con un JSON válido sin markdown:\n"
                        "{\n"
                        '  "destino": "Nombre", "fecha_salida": "Fecha", "duracion": "Tiempo",\n'
                        '  "operador": "Agencia", "precio_base": 0.0, "impuestos_aereo": 0.0,\n'
                        '  "cuotas_cant": 0, "cuota_monto": 0.0, "notas": "Detalles"\n'
                        "}\n"
                        f"Datos extra -> URL: {web_url} | Texto: {email_text}"
                    )
                    
                    messages_content = [{"type": "text", "text": prompt_text}]

                    if uploaded_image:
                        img_bytes = uploaded_image.getvalue()
                        base64_image = base64.b64encode(img_bytes).decode('utf-8')
                        messages_content.append({
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                        })

                    completion = client.chat.completions.create(
                        model=modelo_a_usar,
                        messages=[{"role": "user", "content": messages_content}],
                        temperature=0.1,
                        response_format={"type": "json_object"}
                    )

                    datos_extraidos = json.loads(completion.choices[0].message.content)
                    for k, v in datos_extraidos.items():
                        if v:
                            st.session_state["datos_viaje"][k] = v

                    st.session_state["form_version"] += 1
                    st.rerun()

                except Exception as e:
                    st.error(f"Error con {modelo_a_usar}: {str(e)}")

    st.markdown("---")
    st.subheader("2. Resumen y Edición")
    v = st.session_state["form_version"]

    with st.form("form_presupuesto"):
        c1, c2, c3, c4 = st.columns(4)
        destino = c1.text_input("Destino", value=st.session_state["datos_viaje"]["destino"], key=f"d_{v}")
        fecha_salida = c2.text_input("Salida", value=st.session_state["datos_viaje"]["fecha_salida"], key=f"f_{v}")
        duracion = c3.text_input("Duración", value=st.session_state["datos_viaje"]["duracion"], key=f"dur_{v}")
        operador = c4.text_input("Operador", value=st.session_state["datos_viaje"]["operador"], key=f"o_{v}")

        col_a, col_b, col_c, col_d = st.columns(4)
        precio_base = col_a.number_input("Base (USD)", value=float(st.session_state["datos_viaje"]["precio_base"]), key=f"pb_{v}")
        impuestos_aereo = col_b.number_input("Imp. Aéreos (USD)", value=float(st.session_state["datos_viaje"]["impuestos_aereo"]), key=f"ia_{v}")
        cuotas_cant = col_c.number_input("Cuotas", value=int(st.session_state["datos_viaje"]["cuotas_cant"]), key=f"cc_{v}")
        cuota_monto = col_d.number_input("Monto Cuota (USD)", value=float(st.session_state["datos_viaje"]["cuota_monto"]), key=f"cm_{v}")

        col_e, col_f, col_g, col_h = st.columns(4)
        equipaje_extra = col_e.number_input("Equipaje (USD)", value=float(st.session_state["datos_viaje"]["equipaje_extra"]), key=f"eq_{v}")
        comidas_extra = col_f.number_input("Comidas (USD)", value=float(st.session_state["datos_viaje"]["comidas_extra"]), key=f"co_{v}")
        excursiones_extra = col_g.number_input("Tours (USD)", value=float(st.session_state["datos_viaje"]["excursiones_extra"]), key=f"ex_{v}")
        seguro_medico = col_h.number_input("Seguro (USD)", value=float(st.session_state["datos_viaje"]["seguro_medico"]), key=f"se_{v}")

        gastos_personales = st.number_input("Propinas/Otros (USD)", value=float(st.session_state["datos_viaje"]["gastos_personales"]), key=f"gp_{v}")
        notas = st.text_area("Observaciones", value=st.session_state["datos_viaje"]["notas"], key=f"n_{v}")

        btn_calc = st.form_submit_button("🧮 Actualizar Totales")

    datos_actuales = {
        "destino": destino, "fecha_salida": fecha_salida, "duracion": duracion, "operador": operador,
        "precio_base": precio_base, "impuestos_aereo": impuestos_aereo, "cuotas_cant": cuotas_cant,
        "cuota_monto": cuota_monto, "equipaje_extra": equipaje_extra, "comidas_extra": comidas_extra,
        "excursiones_extra": excursiones_extra, "seguro_medico": seguro_medico,
        "gastos_personales": gastos_personales, "notas": notas
    }

    total_calculado = (precio_base + impuestos_aereo + equipaje_extra + 
                       comidas_extra + excursiones_extra + seguro_medico + gastos_personales)

    st.metric("Total Estimado del Viaje", f"USD ${total_calculado:,.2f}")

    if st.button("💾 Guardar Presupuesto"):
        guardar_en_db(datos_actuales, total_calculado)
        st.success("¡Guardado en base de datos!")

    col_dl1, col_dl2 = st.columns(2)
    col_dl1.download_button("📊 Descargar Excel", generar_excel(datos_actuales, total_calculado), f"{destino}.xlsx")
    col_dl2.download_button("📄 Descargar PDF", generar_pdf(datos_actuales, total_calculado), f"{destino}.pdf")

elif menu == "📜 Historial Guardado":
    st.subheader("📜 Presupuestos Guardados")
    st.dataframe(cargar_historial(), use_container_width=True)